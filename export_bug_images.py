import os
import configparser
import requests
from openpyxl import load_workbook
from pathlib import Path
import re
import filetype
import logging
from logging.handlers import RotatingFileHandler
import tkinter as tk
from tkinter import messagebox


# 加载配置文件
config = configparser.ConfigParser()
config.read('config.ini')

ZENTAO_URL = config['zentao']['url']
USERNAME = config['zentao']['username']
PASSWORD = config['zentao']['password']
EXCEL_FILE = config['excel']['file_path']
BUG_ID_COLUMN = config['excel']['bug_id_column']
START_ROW = int(config['excel']['start_row'])  # 新增读取起始行配置
LOG_FIILE = config['logs']['log_file']

# 配置日志
def setup_logging():
    log_file = LOG_FIILE
    logger.setLevel(logging.INFO)

    # 控制台日志处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(console_formatter)

    # 文件日志处理器
    file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)

    # 添加处理器
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

logger = logging.getLogger(__name__)
setup_logging()

# 登录禅道并返回会话
def login_zentao():
    session = requests.Session()

    # 新登录接口：获取token
    login_url = f"{ZENTAO_URL}/api.php/v1/tokens"
    userinfo = {
        "account": USERNAME,
        "password": PASSWORD
    }
    login_response = session.post(login_url, json=userinfo)
    logger.info("登录状态: %s", login_response.json())
    if login_response.status_code == 201:
        token_data = login_response.json()
        token = token_data.get("token")
        if not token:
            raise Exception("无法获取token")
        # 将token添加到会话头
        session.headers.update({"header": token})
        logger.info("登录成功")
        return session
    else:
        raise Exception("登录失败，请检查用户名或密码")

# 获取指定bugid的图片链接
def get_bug_images(session, bug_id):
    bug_url = f"{ZENTAO_URL}/api.php/v1/bugs/{bug_id}"
    response = session.get(bug_url)
    logger.info("获取bug状态: %s", response.json())
    if response.status_code == 200:
        bug_data = response.json()
        files = bug_data.get("files", {})
        steps = bug_data.get("steps", "")

        # 提取steps中的图片链接
        image_urls = []
        img_tags = re.findall(r'<img.*?src="(.*?)"', steps)
        logger.info("图片链接: %s", img_tags)
        for img_tag in img_tags:
            image_urls.append(img_tag)

        # 提取files中的附件链接
        attachment_urls = {}
        if files:  # 检查files是否为空
            for file_id, file_info in files.items():
                file_title = file_info.get("title", "")  # 获取附件名称
                logger.info(f"附件名称: {file_title}")
                file_url = f"{ZENTAO_URL}{file_info['webPath'].replace("/zentao", "")}"
                attachment_urls[file_title] = file_url  # 使用title作为键

        return {"images": image_urls, "attachments": attachment_urls}
    else:
        raise Exception(f"无法获取bug {bug_id} 的文件信息")

# 导出图片到本地文件夹
def export_images(session, bug_id, file_urls):
    # 修改：仅在有图片或附件时创建文件夹
    if not any(file_urls.values()):
        logger.info(f"Bug {bug_id} 没有附件或图片，跳过文件夹创建")
        return None

    folder_path = Path(f"img/{bug_id}")
    folder_path.mkdir(parents=True, exist_ok=True)  # 创建文件夹

    for file_type, urls in file_urls.items():
        if file_type == "attachments":
            if not file_urls["attachments"]:
                logger.info(f"Bug {bug_id} 没有附件")
                continue
            for title, url in file_urls["attachments"].items():
                file_response = session.get(url)
                if file_response.status_code == 200:
                    file_name = f"{title}"
                    file_path = folder_path / file_name

                    with open(file_path, "wb") as file:
                        file.write(file_response.content)

                    logger.info(f"已保存{file_type}: {file_path}")
        else:
            for idx, url in enumerate(urls):
                file_response = session.get(url)
                if file_response.status_code == 200:
                    file_content = file_response.content
                    file_info = filetype.guess(file_content)
                    file_extension = file_info.extension if file_info else "unknown"
                    file_name = f"{file_type}_{idx + 1}.{file_extension}"
                    file_path = folder_path / file_name

                    with open(file_path, "wb") as file:
                        file.write(file_content)

                    logger.info(f"已保存{file_type}: {file_path}")

    return folder_path

# 更新Excel文件，添加超链接
def update_excel_with_hyperlinks(excel_file, cell, bug_id, folder_path):
    wb = load_workbook(excel_file)
    ws = wb.active
    if int(cell.value) == bug_id:
        # 修改：检查 folder_path 是否有效
        if not folder_path or not Path(folder_path).exists():
            logger.warning(f"警告: 文件夹路径无效或不存在，跳过设置超链接: {folder_path}")
            return

        # 修改：确保 folder_path 是相对路径，并使用 as_posix() 标准化路径
        relative_path = Path(folder_path).as_posix()
        logger.info(f"相对路径: {relative_path}")

        # 修改：修复超链接公式，确保路径部分用双引号包裹
        hyperlink = f'=HYPERLINK("{relative_path}", "{bug_id}")'
        logger.info(f"超链接: {hyperlink}")

        # 设置超链接到单元格
        ws.cell(row=cell.row, column=cell.column).value = hyperlink
    wb.save(excel_file)
    logger.info(f"已更新Excel文件: {excel_file}")

# 主函数
def main():
    session = login_zentao()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    total_rows = ws.max_row - START_ROW + 1
    processed_rows = 0

    # 修改: 明确指定 min_row 和 max_row 参数
    for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
        bug_id = row[0].value
        if bug_id:
            # 新增：清理 bug_id
            cleaned_bug_id = int(bug_id)
            logger.info(f"处理bug ID: {cleaned_bug_id}")
            try:
                image_data = get_bug_images(session, cleaned_bug_id)  # 使用清理后的ID
                if image_data:
                    file_urls = {
                        "images": image_data.get("images", []),
                        "attachments": image_data.get("attachments", {})  # 修改：保持 attachments 的键值对
                    }
                    # 传递 cleaned_bug_id 到 export_images
                    folder_path = export_images(session, cleaned_bug_id, file_urls)
                    update_excel_with_hyperlinks(EXCEL_FILE, row[0], cleaned_bug_id, folder_path)
                else:
                    logger.info(f"Bug {cleaned_bug_id} 没有图片")
            except Exception as e:
                logger.error(f"处理bug {cleaned_bug_id} 时出错: {e}")
        processed_rows += 1
        logger.info(f"进度: {processed_rows}/{total_rows} 行处理完成")

    # 添加导出完成的对话框
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    messagebox.showinfo("导出完成", "所有图片和附件已成功导出！")
    root.destroy()

if __name__ == "__main__":
    main()